import openpyxl

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TrustHub_Test_Records"

# 添加表头（第一行）
headers = [
    "样本名", "输入向量 (in)", "预期输出 (expected_out)",
    "实际输出 (out)", "是否OK", "仿真时间 (ns)",
    "延迟差 (ns)", "功耗估算 (mW)", "备注"
]
for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=header)

# 设置列宽，便于阅读
column_widths = [15, 20, 20, 20, 10, 15, 15, 15, 30]
for col, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# 添加示例数据（第二行，供参考）
example_data = [
    "AES-T100", "8'h00", "8'h63", "8'h63", "OK",
    "100", "0.5", "10.2", "功能验证通过，无异常"
]
for col, data in enumerate(example_data, start=1):
    ws.cell(row=2, column=col, value=data)

# 保存模板文件（工具执行时会生成）
wb.save("TrustHub_Test_Template.xlsx")

# 输出确认
print("Excel模板已生成：TrustHub_Test_Template.xlsx")